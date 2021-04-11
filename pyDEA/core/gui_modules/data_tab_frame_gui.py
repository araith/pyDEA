''' This module contains class DataTabFrame that creates tab with all data.
'''

from openpyxl import load_workbook
import os

from tkinter import StringVar, E, N, W, S, Toplevel
from tkinter.ttk import Frame, Button, LabelFrame
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter.messagebox import askyesno

from pyDEA.core.gui_modules.table_gui import TableFrameWithInputOutputBox
from pyDEA.core.gui_modules.navigation_frame_gui import NavigationForTableFrame
from pyDEA.core.gui_modules.table_modifier_gui import TableModifierFrame
from pyDEA.core.utils.dea_utils import TEXT_FOR_PANEL, FILE_TYPES, center_window
from pyDEA.core.utils.dea_utils import calculate_nb_pages
from pyDEA.core.data_processing.read_data import read_data
from pyDEA.core.gui_modules.load_xlsx_gui import AskSheetName
from pyDEA.core.data_processing.save_data_to_file import save_data_to_xlsx


def remove_star(file_name):
        data_file = file_name
        if file_name:
            if '*' in file_name:
                data_file = data_file[:-1]
        return data_file


class DataTabFrame(Frame):
    ''' This class creates frame with all widgets needed for working with data
        (e.g. loading, saving, modifying, etc.)

        Attributes:
            parent (Tk object): parent of the frame.
            data_from_params_file (StringVar): StringVar object that contains
                name of data file. It is used for communication between this
                frame and parameters frame.
                When parameters are uploaded from file, then file name with
                data might be updated,
                and, hence, this frame must also be updated. If file name is set
                to empty string, then
                this framed is cleared (everything will be deleted). If file
                name points to actual file,
                then data from this file will be uploaded and displayed on
                this frame
            panel (LabelFrame): frame that holds table with data and name of
                the currently loaded data file
            table (TableFrameWithInputOutputBox): object that knows how to
                display data in table and how to modify it
            params_frame (Notebook): frame with parameters. When all data is
                cleared, then we also need to clear parameters
            data (list of list of str or float): stores all loaded or entered
                data without categories,
                for example:
                    >>> [["A", 1, 2, 3], ["B", 1, "", 5], ["C", "", "a", "b"]]
            if_text_modified_str (StringVar): is used for adding star to file
                name if it was modified.
                Technically it is not necessary to store this variable,
                but it is used in unit tests.
            sheet_name (str): name of the sheet in xlsx-file from where data
                has been loaded
            navigation_frame (NavigationForTableFrame): frame that knows how
                to navigate table with data
            table_modifier_frame (TableModifierFrame): frame that adds and
                removes rows and columns of data table

        Args:
            parent (Tk object): parent of this frame
            params_frame (Notebook): frame with parameters. When all data is
                cleared, then we also need to clear parameters
            current_categories (list of str): list with categories, it is
                shared and modified by several objects
            data_from_params_file (StringVar): StringVar object that contains
                name of data file.
                It is used for communication between this frame and
                parameters frame
            str_var_for_input_output_boxes (ObserverStringVar): object that
                contains current input and output categories and notifies
                table with data when data is loaded
                and some categories need to be checked
    '''
    def __init__(self, parent, params_frame, current_categories,
                 data_from_params_file,
                 str_var_for_input_output_boxes, *args, **kw):
        super().__init__(parent, *args, **kw)
        self.parent = parent
        data_from_params_file.trace("w", self.on_params_file_change)
        self.data_from_params_file = data_from_params_file
        self.panel = None
        self.table = None
        self.params_frame = params_frame
        self.data = []
        self.if_text_modified_str = StringVar()
        self.if_text_modified_str.trace('w', self.on_data_modify)
        self.sheet_name = ''
        self.navigation_frame = None
        self.table_modifier_frame = None
        self.create_widgets(current_categories, str_var_for_input_output_boxes)

    def create_widgets(self, current_categories,
                       str_var_for_input_output_boxes):
        ''' Creates widgets on this frame.

            Args:
                current_categories (list of str): list of current categories
                    (might be empty).
                str_var_for_input_output_boxes (ObserverStringVar): object
                    that contains current
                    input and output categories and notifies table with data
                    when data is loaded
                    and some categories need to be checked.
        '''
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        self._create_btns()
        self._create_label_frame()

        table = TableFrameWithInputOutputBox(self.panel, self.params_frame,
                                             self.params_frame.options_frame.
                                             combobox_text_var,
                                             current_categories,
                                             str_var_for_input_output_boxes,
                                             self.if_text_modified_str,
                                             self.data)
        self.table = table
        table.grid(sticky=E+W+S+N, padx=5, pady=5)

        frame_for_navigation_btns = NavigationForTableFrame(self, table)
        table_modifier_frame = TableModifierFrame(self, self.table,
                                                  self.clear_all_on_btn_click,
                                                  frame_for_navigation_btns.
                                                  set_navigation)
        table_modifier_frame.grid(row=1, column=1, sticky=N, pady=15, padx=5)
        self.table_modifier_frame = table_modifier_frame

        frame_for_navigation_btns.grid(row=2, column=0, sticky=E, pady=5,
                                       padx=5)
        self.navigation_frame = frame_for_navigation_btns

    def _create_btns(self):
        ''' Creates load and save buttons.
        '''
        load_button = Button(self, text='Load', command=self.load_file)
        load_button.grid(row=0, column=0, pady=5, padx=5, sticky=W+N)

        frame_for_save_btns = Frame(self)
        save_button = Button(frame_for_save_btns, text='Save',
                             command=self.save_data)
        save_button.grid(row=0, column=0, pady=2, padx=5, sticky=E+N)
        save_as_button = Button(frame_for_save_btns, text='Save as...',
                                command=self.save_data_as)
        save_as_button.grid(row=1, column=0, pady=2, padx=5, sticky=E+N)
        frame_for_save_btns.grid(row=0, column=1, pady=5, padx=5, sticky=E+N)

    def _create_label_frame(self):
        ''' Creates label frame.
        '''
        self.panel = panel = LabelFrame(self, text=TEXT_FOR_PANEL)
        panel.columnconfigure(0, weight=1)
        panel.rowconfigure(0, weight=1)
        panel.grid(row=1, column=0,
                   padx=5, pady=5, sticky=E+W+S+N)

    def get_data_file_name(self):
        ''' Returns name of a file with data if it was loaded.

            File name is stored in label frame.
        '''
        data_file = self.panel.cget('text')
        nb_letters = len(TEXT_FOR_PANEL)
        return data_file[nb_letters:]

    def remove_star_from_panel(self):
        ''' Removes star from data file name if it was there.
        '''
        data_file = self.get_data_file_name()
        data_file = remove_star(data_file)
        self.panel.config(text=TEXT_FOR_PANEL + data_file)

    def save_data(self):
        ''' Saves data to an existing or new file.

            This method is called when user presses "Save" button.
            If user modified open data file, then modifications will
            be saved in this file.
            If user typed data, this method redirects user to
            save_data_as() method.
        '''
        data_file = self.get_data_file_name()
        if data_file == '*':  # when file was created and modified
            self.save_data_as()
        elif data_file:
            self.save_data_to_given_file(remove_star(data_file),
                                         sheet_name=self.sheet_name)
            self.remove_star_from_panel()

    def save_data_as(self):
        ''' Saves data to a new file.

            Calls dialogue to ask user where to save file and saves
            data file there. Default extension is .xlsx.
        '''
        file_name = self._call_file_save_as_dialog()
        if file_name:
            self.save_data_to_given_file(file_name)
            self.panel.config(text=TEXT_FOR_PANEL + file_name)

    def _call_file_save_as_dialog(self):
        ''' Calls save as dialogue.

            This method is redefined in unit tests.
        '''
        return asksaveasfilename(filetypes=FILE_TYPES, defaultextension='.xlsx')

    def save_data_to_given_file(self, data_file, sheet_name='Data'):
        ''' Saves data to a given file.

            Args:
                data_file (str): file name where data should be stored.
                sheet_name (str): sheet name where data should be stored.
                Defaults to "Data".
        '''
        categories = [self.table.cells[0][col].get() for col
                      in range(self.table.nb_cols)]
        save_data_to_xlsx(data_file, categories, self.data, sheet_name)

    def on_data_modify(self, *args):
        ''' Adds star to label frame when data was modified.

            Args:
                *args: are not used in this method, but are provided by
                    StringVar trace routine.
        '''
        data_file = self.panel.cget('text')
        if data_file and '*' not in data_file:
            self.panel.config(text=data_file + '*')

    def load_file(self):
        ''' Asks user which data file should be loaded and loads specified file.

            Only xlsx and csv files are allowed.
        '''
        file_name = self._call_open_file_dialogue()
        if file_name:
            self.show_loaded_data(file_name)

    def _call_open_file_dialogue(self):
        ''' Calls open file dialogue. This method is overridden in unit tests.

            Returns:
                (str): file name of file that should be loaded or empty string
                    if the user pressed cancel
        '''
        return askopenfilename(title='Choose a file', filetypes=FILE_TYPES)

    def clear_all_on_btn_click(self):
        ''' Clears all data if user agrees.

            Asks the user if he/she wants to clear all data, clears all data
            if user agrees.
            This method is called when user presses button "Clear all".
        '''
        if (self.data and
            askyesno("Verify",
                     "Are sure you want to delete all input data?")):
            self.clear_all()

    def clear_all(self):
        ''' Clears all data.
        '''
        self.table.clear_all_data()
        self.table.deselect_all_boxes()
        self.params_frame.clear_all()
        self.panel.config(text=TEXT_FOR_PANEL)
        self.parent.reset_progress_bar()
        self.sheet_name = ''
        self.navigation_frame.reset_navigation()

    def on_params_file_change(self, *args):
        ''' Loads data or clears all after parameters are loaded from file.

            If file with parameters contains data file, then this data
            will be attempted to load.
            If file with parameters doesn't contain data file
            (or user specified to load parameters
            without data), then all previously loaded data will remain.

            Args:
                *args: are not used in this methods and are provided by
                    trace method of StringVar
        '''
        params_file = self.data_from_params_file.get()
        if params_file:
            self.show_loaded_data(params_file)

    def show_loaded_data(self, file_name):
        ''' Loads data file.

            This method asks user from what sheet data must be loaded if
            there are more then 1 sheet
            in data file. Changes name of data file in label frame and
            calls method that displays
            data in the table.

            Args:
                file_name (str): data file
        '''
        just_name, extension = os.path.splitext(file_name)
        should_proceed = False
        if extension == '.xlsx':
            book = load_workbook(file_name, data_only = True)
            names = book.sheetnames
            nb_names = len(names)
            if nb_names == 0:
                return
            elif nb_names > 1:
                self.sheet_name = self.call_dialog(names)
            else:
                self.sheet_name = names[0]
            if self.sheet_name:
                should_proceed = True
        else:
            self.sheet_name = ''  # csv does not support sheets
            should_proceed = True
        if should_proceed:
            sheet_name_copy = self.sheet_name
            self.clear_all()
            # set data file name in parameters
            # (this field is cleared in clear_all())
            self.params_frame.params.update_parameter('DATA_FILE', file_name)
            self.sheet_name = sheet_name_copy
            self.panel.config(text=TEXT_FOR_PANEL + file_name)
            categories, coefficients, dmu_name, self.sheet_name = read_data(
                file_name, self.sheet_name)
            self.show_data(categories, coefficients, dmu_name)
            self.remove_star_from_panel()

    def call_dialog(self, names):
        ''' Calls dialogue that asks user to specify sheet name from where
            data should be loaded.

            Args:
                names (list of str): list of sheet names

            Returns:
                (str): name of the selected sheet if sheet was selected,
                    empty string if user pressed cancel
        '''
        win = Toplevel()
        win.withdraw()
        ask_sheet_name_frame = AskSheetName(win, names)
        var = ask_sheet_name_frame.sheet_name_str
        ask_sheet_name_frame.pack()
        center_window(win)
        win.grab_set()
        win.focus_set()
        win.wait_window()
        return var.get()

    def show_data(self, categories, coefficients, dmu_name):
        ''' Fills table with data.

            Args:
                categories (list of str): list with categories that will
                appear in the first table row
                coefficients (list of list of str or float): data that will
                    appear after categories.
                    Each contained list will be displayed on a new table row
                dmu_name (str): text that appear in the same line with
                    categories before the categories appear
        '''
        self.table.cells[0][0].insert(0, dmu_name)

        nb_needed_cols = len(categories) + 1
        for i in range(self.table.nb_cols, nb_needed_cols):
            self.table.add_column()

        # print categories
        for count, category in enumerate(categories):
            self.table.cells[0][count + 1].insert(0, category)

        # print data
        self.data.clear()
        for value in coefficients:
            self.data.append(value)
        self.table.load_visible_data()

        nb_data_pages = calculate_nb_pages(len(self.data), self.table.nb_rows)

        self.navigation_frame.set_navigation(nb_data_pages)

    def read_coefficients(self):
        ''' Converts coefficients into a proper data structure
            used for running algorithms.

            Returns:
                see read_data for details
        '''
        return self.table.read_coefficients()
