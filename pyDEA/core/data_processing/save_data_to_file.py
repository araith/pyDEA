''' This module implements function to write data to xlsx file.
'''

from openpyxl import load_workbook
import os

from pyDEA.core.data_processing.xlsx_workbook import XlsxWorkbook, XlsxSheet
from pyDEA.core.data_processing.solution_text_writer import OneCsvWriter


def create_workbook(output_file):
    ''' Creates a proper instance of data file writer depending on file extension.
        Supported formats are: xslx and csv.

        Args:
            output_file (str): file name.

        Returns:
            (XlsxWorkbook or OneCsvWriter): created data file
                writer instance.

        Raises:
            ValueError: if a file with not supported extension was provided.
    '''
    if output_file.endswith('.xlsx'):
        work_book = XlsxWorkbook()
    elif output_file.endswith('.csv'):
        work_book = OneCsvWriter(output_file)
    else:
        raise ValueError('File {0} has unsupported output format'.format
                         (output_file))
    return work_book


def save_data_to_xlsx(data_file, categories, data, sheet_name='Data'):
    ''' Writes input data to xlsx-file.

        Args:
            data_file (str): name of new or existing file where data will
                be written.
            categories (list of str): the first line that will be written
                to file, it must contain category names and might also
                contain name of DMUs as the first element.
            data (list of list of int, float or str): list with data,
                for example

                >> data = [['A', 1, 2, 3], ['B', 2, "a", 5], ['C', 0]]
            sheet_name (str, optional): name of the sheet where the data
                will be written, if the existing sheet name is specified,
                data will be overwritten on this sheet, defaults to "Data".

        Raises:
            ValueError: if sheet_name is empty string or None, or any
                empty value for xlsx files.
    '''
    if data_file.endswith('.xlsx') and not sheet_name:
        raise ValueError('Sheet name is not specified')

    if os.path.isfile(data_file):
        work_book = create_workbook(data_file)
        if data_file.endswith('.xlsx'):
            # copy data from all sheets except the one the user modified if any
            # copy old file and clear sheet to be overwritten with new data
            work_book.workbook = load_workbook(data_file, data_only = True)
            #due to data_only the formulas are replaced by numbers, but otherwise 
            #the file would get useless for pyDEA until opening it again in Excel and recalculating the equations
            work_book.nb_sheets = len(work_book.workbook.sheetnames)
            if sheet_name in work_book.workbook.sheetnames:
                ws = work_book.workbook[sheet_name]
                ws.delete_rows(1, amount = ws.max_row)
                work_sheet = XlsxSheet(ws)
            else:
                work_sheet = work_book.add_sheet(sheet_name)
        else:
            work_sheet = work_book.add_sheet(sheet_name)
    else:
        work_book = create_workbook(data_file)
        work_sheet = work_book.add_sheet(sheet_name)

    # save categories
    for count, category in enumerate(categories):
        work_sheet.write(0, count, category.strip())

    # save data
    row_index = 1
    for values in data:
        for col, coeff in enumerate(values):
            try:
                val = float(coeff)
            except ValueError:
                val = coeff
            work_sheet.write(row_index, col, val)
        row_index += 1
    work_book.save(data_file)
