''' This module contains classes responsible for wrapping openpyxl functionality
    for creating xlsx files.
'''
from openpyxl import Workbook


class XlsxSheet(object):
    ''' This class wraps openpyxl work sheet so that it can be used by xls writer
        class.

        Attributes:
            name (str): work sheet name.
            worksheet (openpyxl work sheet): worksheet to wrap.

        Args:
            worksheet (openpyxl work sheet): worksheet to wrap.
    '''
    def __init__(self, worksheet):
        self.worksheet = worksheet

    @property
    def name(self):
        return self.worksheet.title

    @name.setter
    def name(self, value):
        self.worksheet.title = value

    def write(self, row, col, value):
        ''' Writes given value to a cell with specified row and column.

            Args:
                row (int): row index.
                col (int): column index.
                value (object): value to write.
        '''
        cell = self.worksheet.cell(row=row, column=col)
        cell.value = value


class XlsxWorkbook(object):
    ''' This class wraps openpyxl workbook so it can be used by
        xls writer class.

        Attributes:
            workbook (openpyxl workbook): workbook.
            nb_sheets (int): number of added work sheets.
    '''
    def __init__(self):
        self.workbook = Workbook()
        self.nb_sheets = 0

    def add_sheet(self, sheet_name):
        ''' Adds one sheet to the workbook.

            Args:
                sheet_name (str): name of the work sheet.

            Returns:
                XlsxSheet: created work sheet.
        '''
        self.nb_sheets += 1
        if self.nb_sheets == 1:
            worksheet = self.workbook.active
        else:
            worksheet = self.workbook.create_sheet()
        worksheet.title = sheet_name
        return XlsxSheet(worksheet)

    def save(self, file_name):
        ''' Saves workbook to a given file.

            Args:
                file_name (str): name of the file where workbook should be
                    saved.
        '''
        self.workbook.save(file_name)
