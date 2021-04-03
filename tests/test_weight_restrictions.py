import pytest
from openpyxl import Workbook
import datetime

from pyDEA.core.data_processing.read_data_from_xls import read_data
from pyDEA.core.data_processing.read_data_from_xls import construct_input_data_instance
from pyDEA.core.data_processing.read_data_from_xls import validate_data, convert_to_dictionary
from pyDEA.core.models.multiplier_model_base import MultiplierModelBase
from pyDEA.core.models.multiplier_model import MultiplierInputOrientedModel
from pyDEA.core.models.multiplier_model_decorators import MultiplierModelWithAbsoluteWeightRestrictions
from pyDEA.core.models.multiplier_model_decorators import MultiplierModelWithVirtualWeightRestrictions
from pyDEA.core.models.multiplier_model_decorators import MultiplierModelWithPriceRatioConstraints
from pyDEA.core.models.envelopment_model_decorators import EnvelopmentModelWithAbsoluteWeightRestrictions
from pyDEA.core.models.envelopment_model_decorators import EnvelopmentModelWithVirtualWeightRestrictions
from pyDEA.core.models.envelopment_model_decorators import EnvelopmentModelWithPriceRatioConstraints
from pyDEA.core.models.bound_generators import generate_upper_bound_for_efficiency_score
from pyDEA.core.data_processing.write_data_to_xls import XLSWriter
from pyDEA.core.data_processing.parameters import Parameters
from pyDEA.core.utils.dea_utils import clean_up_pickled_files
from pyDEA.core.data_processing.parameters import parse_parameters_from_file
from pyDEA.core.utils.model_builder import build_models
from pyDEA.core.models.envelopment_model_base import EnvelopmentModelBase
from pyDEA.core.models.envelopment_model import EnvelopmentModelInputOriented
from pyDEA.core.models.bound_generators import generate_upper_bound_for_efficiency_score
from pyDEA.core.models.envelopment_model_decorators import DefaultConstraintCreator

import tests.utils_for_tests as utils_for_tests
from tests.utils_for_tests import DEA_example_data


@pytest.fixture
def data(request, DEA_example_data):
    data = DEA_example_data
    data.add_input_category('I1')
    data.add_input_category('I2')
    data.add_output_category('O1')
    data.add_output_category('O2')
    return data


def test_abs_weight_restrictions_multiplier_model(data):

    base_model = MultiplierModelBase(data, 0,
                                     MultiplierInputOrientedModel())

    bounds = {'I2': (0.01, 0.5)}
    model = MultiplierModelWithAbsoluteWeightRestrictions(base_model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_abs_weights_multi_output.xls')

    bounds = {'I2': (None, 0.05)}
    model = MultiplierModelWithAbsoluteWeightRestrictions(base_model, bounds)

    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()

    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, bounds)

    work_book2 = Workbook()
    writer = XLSWriter(Parameters(), work_book2, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book2.save('tests/test_abs_weights_upper_bound_multi_output.xls')


def test_virtual_weight_restrictions_multiplier_model(data):

    model = MultiplierModelBase(data, 0,
                                MultiplierInputOrientedModel())

    bounds = {'I1': (None, 0.5)}
    model = MultiplierModelWithVirtualWeightRestrictions(model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_virtual_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_virtual_weights_multi_output.xls')


def test_abs_and_virtual_restrictions_multiplier_model(data):

    model = MultiplierModelBase(data, 0,
                                MultiplierInputOrientedModel())

    bounds = {'I1': (None, 0.4)}
    model = MultiplierModelWithVirtualWeightRestrictions(model, bounds)

    abs_bounds = {'I2': (None, 0.2)}
    model = MultiplierModelWithAbsoluteWeightRestrictions(model, abs_bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, abs_bounds)
    utils_for_tests.check_if_category_is_within_virtual_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_virtual_and_abs_weights_multi_output.xls')


def test_price_ratio_multiplier_model(data):

    model = MultiplierModelBase(data, 0,
                                MultiplierInputOrientedModel())

    bounds = {('I1', 'I2'): (None, 0.4), ('O1', 'O2'): (0.01, None)}
    model = MultiplierModelWithPriceRatioConstraints(model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_price_ratio_constraints(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_price_ratio_multi_output.xls')


def test_all_constraints_multiplier_model(data):

    model = MultiplierModelBase(data, 0,
                                MultiplierInputOrientedModel())

    bounds = {'I1': (None, 0.4)}
    model = MultiplierModelWithVirtualWeightRestrictions(model, bounds)

    abs_bounds = {'I2': (None, 0.2)}
    model = MultiplierModelWithAbsoluteWeightRestrictions(model, abs_bounds)

    ratio_bounds = {('I1', 'I2'): (None, 0.4), ('O1', 'O2'): (0.01, None)}
    model = MultiplierModelWithPriceRatioConstraints(model, ratio_bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, abs_bounds)
    utils_for_tests.check_if_category_is_within_virtual_limits(
        model_solution, bounds)
    utils_for_tests.check_if_category_is_within_price_ratio_constraints(
        model_solution, ratio_bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_all_constraints_multi_output.xls')


def test_abs_restrictions_env_model(data):
    model = EnvelopmentModelBase(data,
                                 EnvelopmentModelInputOriented(
                                     generate_upper_bound_for_efficiency_score),
                                 DefaultConstraintCreator())

    bounds = {'I2': (0.01, 0.5)}
    model = EnvelopmentModelWithAbsoluteWeightRestrictions(model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_abs_constraints_env_output.xls')


def test_abs_restrictions_env_model_output(data):
    filename = 'tests/DEA_Harish_parameters.txt'
    params = parse_parameters_from_file(filename)
    categories, data, dmu_name, sheet_name = read_data(
        params.get_parameter_value('DATA_FILE'))
    coefficients, has_same_dmus = convert_to_dictionary(data)
    model_input = construct_input_data_instance(categories, coefficients)
    models, all_params = build_models(params, model_input)
    assert len(models) == 1 and len(all_params) == 1
    model = models[0]

    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    bounds = {'Urban Roads (%)': (None, 0.003)}
    utils_for_tests.check_if_category_is_within_abs_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_abs_constraints_env_outoriented_output.xls')


def test_virual_restrictions_env_model(data):
    model = EnvelopmentModelBase(data,
                                 EnvelopmentModelInputOriented(
                                     generate_upper_bound_for_efficiency_score),
                                 DefaultConstraintCreator())

    bounds = {'I2': (0.01, 0.5)}
    model = EnvelopmentModelWithVirtualWeightRestrictions(model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_virtual_limits(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_virtual_constraints_env_output.xls')


def test_price_ratio_restrictions_env_model(data):
    model = EnvelopmentModelBase(data,
                                 EnvelopmentModelInputOriented(
                                     generate_upper_bound_for_efficiency_score),
                                 DefaultConstraintCreator())

    bounds = {('I1', 'I2'): (None, 0.4), ('O1', 'O2'): (0.01, None)}
    model = EnvelopmentModelWithPriceRatioConstraints(model, bounds)
    start_time = datetime.datetime.now()
    model_solution = model.run()
    end_time = datetime.datetime.now()
    utils_for_tests.check_if_category_is_within_price_ratio_constraints(
        model_solution, bounds)

    work_book = Workbook()
    writer = XLSWriter(Parameters(), work_book, datetime.datetime.today(),
                       (end_time - start_time).total_seconds())
    writer.write_data(model_solution)
    work_book.save('tests/test_price_ratio_env_output.xls')


def test_price_ratio_restrictions_medium_env_model():
    categories, data, dmu_name, sheet_name = read_data(
        'tests/dataFromDEAbook_page181.xls')
    coefficients, has_same_dmus = convert_to_dictionary(data)
    assert has_same_dmus is False
    assert validate_data(categories, coefficients) is True
    data = construct_input_data_instance(categories, coefficients)
    print(data.categories)
    data.add_input_category('Doctors')
    data.add_input_category('Nurses')
    data.add_output_category('Outpatients')
    data.add_output_category('Inpatients')

    model = MultiplierModelBase(data, 0,
                                MultiplierInputOrientedModel())

    ratio_bounds = {('Nurses', 'Doctors'): (0.2, 5),
                    ('Inpatients', 'Outpatients'): (0.2, 5)}
    model = MultiplierModelWithPriceRatioConstraints(model, ratio_bounds)

    model_solution = model.run()
    utils_for_tests.check_if_category_is_within_price_ratio_constraints(
        model_solution, ratio_bounds)

    utils_for_tests.check_optimal_solution_status_and_sizes(
        model_solution, data)
    dmus = ['H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7',
            'H8', 'H9', 'H10', 'H11', 'H12', 'H13', 'H14']
    utils_for_tests.check_efficiency_scores(dmus, [0.926, 1, 1, 0.634, 0.82, 1,
                                                   0.803, 0.872, 0.982, 1,
                                                   0.849,
                                                   0.93, 0.74, 0.929],
                                            model_solution, data, 1e-3)
    clean_up_pickled_files()
